from __future__ import annotations

import json
import os
import re
import shutil
import sys
import tempfile
import threading
import time
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Generator, Literal
from urllib.parse import unquote

from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from openpyxl import load_workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.formula.tokenizer import Token, Tokenizer
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from pydantic import BaseModel, Field

APP_ROOT = Path(__file__).resolve().parents[1]
if str(APP_ROOT) not in sys.path:
    sys.path.append(str(APP_ROOT))

from config_loader import get_config_value
from llm_client import LLMClient


app = FastAPI(title="FormulaLens")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:8080"],
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type"],
)

UPLOADS_DIR = Path(__file__).resolve().parent / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
REGISTRY_PATH = UPLOADS_DIR / "registry.json"
PROMPTS_DIR = APP_ROOT / "prompts"
MAX_UPLOAD_SIZE = 200_000_000
DEFAULT_MODEL = get_config_value("SCRIPT_JUDGE_MODEL") or "gpt-4.1-2025-04-14"
MAX_TRACE_DEPTH = 5
FILE_ID_RE = re.compile(r"^[0-9a-f]{12}$")
CELL_RE = re.compile(r"^([A-Z]{1,3})(\d{1,7})$")
RANGE_TOKEN_RE = re.compile(
    r"^(?:(?P<sheet>'[^']+'|[^!]+)!)?(?P<ref>\$?[A-Z]{1,3}\$?\d{1,7}(?::\$?[A-Z]{1,3}\$?\d{1,7})?)$"
)
INLINE_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z_][\w .-]*)!)?(?P<cell>\$?[A-Z]{1,3}\$?\d{1,7})"
)

store: dict[str, dict[str, Any]] = {}
ref_index_cache: dict[str, dict[str, list[tuple[str, str]]]] = {}
sheet_cache: dict[str, dict[str, Any]] = {}
tables_cache: dict[str, list[dict[str, Any]]] = {}
explanation_cache: dict[str, str] = {}
_referenced_cache: dict[str, set[str]] = {}
task_store: dict[str, "LLMTask"] = {}
task_by_cache_key: dict[str, str] = {}
_file_locks: dict[str, threading.Lock] = {}

_registry_ready = threading.Event()
_task_lock = threading.Lock()
_file_locks_lock = threading.Lock()
_boot_status = {
    "stage": "booting",
    "detail": "Starting FormulaLens",
    "files_total": 0,
    "files_loaded": 0,
    "ready": False,
}


class TraceReq(BaseModel):
    file_id: str
    sheet: str
    cell: str
    max_depth: int = 5


class TableTraceReq(BaseModel):
    file_id: str
    sheet: str
    range: str
    max_depth: int = 5


class SaveTablesReq(BaseModel):
    tables: list[dict]


class ExplainReq(BaseModel):
    trace: dict
    file_id: str = ""
    sheet: str = ""
    cell: str = ""
    regenerate: bool = False
    model: str = DEFAULT_MODEL


class BatchExplainReq(BaseModel):
    metrics: list[dict]
    model: str = DEFAULT_MODEL


class OptimizeReq(BaseModel):
    trace: dict
    label: str = ""
    model: str = DEFAULT_MODEL


class ChatReq(BaseModel):
    file_id: str
    message: str
    sheet: str = ""
    selected_tables: list[str] = Field(default_factory=list)
    history: list[dict] = Field(default_factory=list)
    model: str = DEFAULT_MODEL
    mode: str = "auto"
    focus_cells: list[str] = Field(default_factory=list)


class CellEdit(BaseModel):
    cell: str
    value: Any = None
    formula: str | None = None


class EditCellsReq(BaseModel):
    file_id: str
    sheet: str
    edits: list[CellEdit]


class CellFormat(BaseModel):
    fill: str | None = None
    font_color: str | None = None
    bold: bool | None = None
    italic: bool | None = None
    number_format: str | None = None


class FormatCellsReq(BaseModel):
    file_id: str
    sheet: str
    cells: list[str]
    format: CellFormat


class InsertChartReq(BaseModel):
    file_id: str
    sheet: str
    chart_spec: dict
    near_range: str | None = None


class SummaryExplainReq(BaseModel):
    file_id: str
    metrics: list[dict]
    regenerate: bool = False
    model: str = DEFAULT_MODEL


@dataclass
class LLMTask:
    task_id: str
    task_type: str
    cache_key: str
    file_id: str
    status: Literal["running", "done", "error", "cancelled"] = "running"
    chunks: list[str] = field(default_factory=list)
    full_text: str = ""
    error_msg: str = ""
    cancel_flag: threading.Event = field(default_factory=threading.Event)
    created_at: float = field(default_factory=time.time)


_INJECTION_KEYWORDS = (
    "ignore previous",
    "ignore all prior",
    "system prompt",
    "developer prompt",
    "jailbreak",
    "override instructions",
    "reveal prompt",
)
_INJECTION_TARGETS = ("prompt", "system", "secret", "hidden instructions")
_HARM_PHRASES = (
    re.compile(r"\b(exfiltrate|steal|dump)\b.*\b(credentials|passwords|tokens)\b", re.I),
    re.compile(r"\bmalware\b|\bransomware\b|\bkeylogger\b", re.I),
)
_PERSONA_DATA_HINTS = (
    "trend", "distribution", "outlier", "average", "median", "stddev", "correlation",
    "yoy", "year over year", "growth rate", "anomaly", "spike", "dip", "histogram",
    "p-value", "regression", "forecast", "seasonality", "data analy",
)
_PERSONA_BUSINESS_HINTS = (
    "summary", "executive", "review", "stakeholder", "exec ", "leadership",
    "narrative", "story", "talking points", "presentation", "qbr", "earnings",
    "board", "memo", "headline", "tldr",
)
_PERSONA_EXCEL_HINTS = (
    "formula", "sumifs", "vlookup", "xlookup", "index", "match", "named range",
    "cell", "sheet", "workbook", "pivot", "rewrite", "optimi",
)


def _sse(payload: dict[str, Any]) -> str:
    return f"data: {json.dumps(payload)}\n\n"


def _sanitize_filename(filename: str) -> str:
    name = Path(filename or "workbook.xlsx").name
    return re.sub(r"[^A-Za-z0-9._ -]+", "_", name)


def _atomic_write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(payload, indent=2))
    tmp.replace(path)


def _get_file_lock(fid: str) -> threading.Lock:
    with _file_locks_lock:
        if fid not in _file_locks:
            _file_locks[fid] = threading.Lock()
        return _file_locks[fid]


def _registry_data() -> dict[str, dict[str, Any]]:
    return {
        fid: {
            "filename": entry["filename"],
            "path": entry["path"],
            "sheets": entry["sheets"],
        }
        for fid, entry in store.items()
    }


def _save_registry() -> None:
    _atomic_write_json(REGISTRY_PATH, _registry_data())


def _explanations_path(fid: str) -> Path:
    return UPLOADS_DIR / fid / "explanations.json"


def _load_explanations(fid: str) -> None:
    path = _explanations_path(fid)
    if not path.exists():
        return
    try:
        data = json.loads(path.read_text())
    except Exception:
        return
    for cache_key, text in data.items():
        if isinstance(text, str):
            explanation_cache[f"{fid}:{cache_key}"] = text


def _save_explanations(fid: str) -> None:
    prefix = f"{fid}:"
    payload = {key[len(prefix):]: value for key, value in explanation_cache.items() if key.startswith(prefix)}
    _atomic_write_json(_explanations_path(fid), payload)


def _clear_file_caches(fid: str, sheet: str | None = None) -> int:
    cleared = 0
    keys = list(sheet_cache)
    table_keys = list(tables_cache)
    if sheet:
        prefix = f"{fid}:{sheet}"
        for key in keys:
            if key == prefix:
                sheet_cache.pop(key, None)
                cleared += 1
        for key in table_keys:
            if key == prefix:
                tables_cache.pop(key, None)
                cleared += 1
    else:
        for key in keys:
            if key.startswith(f"{fid}:"):
                sheet_cache.pop(key, None)
                cleared += 1
        for key in table_keys:
            if key.startswith(f"{fid}:"):
                tables_cache.pop(key, None)
                cleared += 1
    ref_index_cache.pop(fid, None)
    _referenced_cache.pop(fid, None)
    return cleared


def _reload_file_from_disk(fid: str) -> None:
    entry = _file_or_404(fid)
    path = Path(entry["path"])
    wb_f = load_workbook(path, data_only=False)
    wb_v = load_workbook(path, data_only=True)
    store[fid] = {
        "path": str(path),
        "filename": entry["filename"],
        "wb_f": wb_f,
        "wb_v": wb_v,
        "sheets": wb_f.sheetnames,
    }
    _build_ref_index(fid, force=True)


def _file_or_404(fid: str) -> dict[str, Any]:
    entry = store.get(fid)
    if not entry:
        raise HTTPException(status_code=404, detail="file not found")
    return entry


def _sheet_or_404(entry: dict[str, Any], sheet: str):
    sheet_name = unquote(sheet)
    if sheet_name not in entry["wb_f"].sheetnames:
        raise HTTPException(status_code=404, detail="sheet not found")
    return sheet_name, entry["wb_f"][sheet_name], entry["wb_v"][sheet_name]


def _parse_cell_ref(ref: str) -> tuple[int, int] | None:
    match = CELL_RE.fullmatch(ref.replace("$", "").upper())
    if not match:
        return None
    col_letters, row_text = match.groups()
    col_num = 0
    for ch in col_letters:
        col_num = col_num * 26 + (ord(ch) - 64)
    return col_num, int(row_text)


def _col_to_letter(n: int) -> str:
    return get_column_letter(n)


def parse_refs(formula: str | None, default_sheet: str) -> tuple[list[tuple[str, str]], list[tuple[str, str]], bool]:
    if not formula:
        return [], [], False
    cell_refs: list[tuple[str, str]] = []
    range_refs: list[tuple[str, str]] = []
    has_external = False
    try:
        tokens = Tokenizer(formula).items
    except Exception:
        tokens = []
    for token in tokens:
        if token.type != Token.OPERAND or token.subtype != Token.RANGE:
            continue
        raw = token.value.replace("$", "")
        if raw.startswith("["):
            has_external = True
            raw = re.sub(r"^\[[^\]]+\]", "", raw)
        match = RANGE_TOKEN_RE.match(raw)
        if not match:
            continue
        sheet_name = (match.group("sheet") or default_sheet).strip("'")
        ref = match.group("ref").upper()
        if ":" in ref:
            range_refs.append((sheet_name, ref))
        else:
            cell_refs.append((sheet_name, ref))
    return list(dict.fromkeys(cell_refs)), list(dict.fromkeys(range_refs)), has_external


def _cell_formula(ws_f, cell_ref: str) -> str | None:
    try:
        raw = ws_f[cell_ref].value
    except Exception:
        return None
    return raw if isinstance(raw, str) and raw.startswith("=") else None


def _cell_value(ws_v, cell_ref: str) -> str:
    try:
        raw = ws_v[cell_ref].value
    except Exception:
        return "[bad ref]"
    return "" if raw is None else str(raw)


def _cell_has_formula(ws_f, cell_ref: str) -> bool:
    return _cell_formula(ws_f, cell_ref) is not None


def _argb_to_hex(argb: str | None) -> str | None:
    if not argb:
        return None
    argb = str(argb).upper()
    if len(argb) == 8:
        alpha = argb[:2]
        rgb = argb[2:]
        if alpha == "00":
            return None
        return f"#{rgb}"
    if len(argb) == 6:
        return f"#{argb}"
    return None


def _extract_cell_style(cell) -> dict[str, Any] | None:
    out: dict[str, Any] = {}
    fill = getattr(cell.fill, "fgColor", None)
    fill_hex = _argb_to_hex(getattr(fill, "rgb", None))
    if fill_hex and fill_hex != "#FFFFFF":
        out["bg"] = fill_hex
    font_color = _argb_to_hex(getattr(getattr(cell.font, "color", None), "rgb", None))
    if font_color and font_color not in {"#000000", "#1F2937"}:
        out["fg"] = font_color
    if cell.font and cell.font.bold:
        out["b"] = 1
    if cell.font and cell.font.italic:
        out["i"] = 1
    if cell.number_format and cell.number_format not in {"General", "@"}:
        out["nf"] = cell.number_format
    return out or None


def _is_numericish(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True
    if not isinstance(value, str):
        return False
    text = value.strip().replace(",", "").replace("%", "").replace("$", "")
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def _compute_cell_label(entry: dict[str, Any], sheet: str, cell_ref: str) -> str:
    ws = entry["wb_v"][sheet]
    parsed = _parse_cell_ref(cell_ref)
    if not parsed:
        return f"{sheet}!{cell_ref}"
    col, row = parsed
    left_labels: list[str] = []
    saw_left_text = False
    for c in range(col - 1, 0, -1):
        raw = ws.cell(row=row, column=c).value
        if isinstance(raw, str) and raw.strip() and not raw.startswith("="):
            if not _is_numericish(raw):
                left_labels.append(raw.strip())
                saw_left_text = True
                continue
        if saw_left_text and raw not in (None, ""):
            break
    left_labels.reverse()
    top_labels: list[str] = []
    saw_top_text = False
    for r in range(row - 1, 0, -1):
        raw = ws.cell(row=r, column=col).value
        if isinstance(raw, str) and raw.strip() and not raw.startswith("="):
            if not _is_numericish(raw):
                top_labels.append(raw.strip())
                saw_top_text = True
                continue
        if saw_top_text and raw not in (None, ""):
            break
    top_labels.reverse()
    labels = [label for label in [*left_labels, *top_labels] if label]
    return " · ".join(labels) if labels else f"{sheet}!{cell_ref}"


def _get_cell_label(entry: dict[str, Any], sheet: str, cell_ref: str) -> str:
    return _compute_cell_label(entry, sheet, cell_ref)


def _add_meta_to_trace(entry: dict[str, Any], node: dict[str, Any]) -> dict[str, Any]:
    node["meta"] = _get_cell_label(entry, node["sheet"], node["cell"])
    for dep in node.get("deps", []):
        _add_meta_to_trace(entry, dep)
    return node


def _get_range_headers(entry: dict[str, Any], sheet: str, range_ref: str) -> list[str]:
    if sheet not in entry["wb_v"].sheetnames:
        return []
    try:
        min_col, min_row, max_col, _ = range_boundaries(range_ref)
    except Exception:
        return []
    ws = entry["wb_v"][sheet]
    headers: list[str] = []
    header_row = max(min_row - 1, 1)
    for col in range(min_col, max_col + 1):
        found = None
        for r in range(header_row, 0, -1):
            raw = ws.cell(row=r, column=col).value
            if isinstance(raw, str) and raw.strip() and not _is_numericish(raw) and not raw.startswith("="):
                found = raw.strip()
                break
        headers.append(found or _col_to_letter(col))
    return headers


def _cell_in_range(cell: str, range_ref: str) -> bool:
    parsed = _parse_cell_ref(cell)
    if not parsed:
        return False
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except Exception:
        return False
    col, row = parsed
    return min_col <= col <= max_col and min_row <= row <= max_row


def _expand_range_formula_cells(entry: dict[str, Any], sheet: str, range_ref: str) -> list[str]:
    if sheet not in entry["wb_f"].sheetnames:
        return []
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    except Exception:
        return []
    ws = entry["wb_f"][sheet]
    out: list[str] = []
    scanned = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            coord = f"{_col_to_letter(col)}{row}"
            if _cell_has_formula(ws, coord):
                out.append(coord)
            scanned += 1
            if scanned >= 10_000:
                return out
    return out


def _trace_node(entry: dict[str, Any], sheet: str, cell: str, path: set[str] | None = None, depth: int = 0, max_depth: int = MAX_TRACE_DEPTH) -> dict[str, Any]:
    path = path or set()
    key = f"{sheet}!{cell}"
    if key in path:
        return {"cell": cell, "sheet": sheet, "value": "[circular]", "formula": None, "deps": [], "ranges": [], "external": False}
    if sheet not in entry["wb_f"].sheetnames:
        return {"cell": cell, "sheet": sheet, "value": f"[sheet '{sheet}' not found]", "formula": None, "deps": [], "ranges": [], "external": False}
    ws_f = entry["wb_f"][sheet]
    ws_v = entry["wb_v"][sheet]
    formula = _cell_formula(ws_f, cell)
    value = _cell_value(ws_v, cell)
    node = {
        "cell": cell,
        "sheet": sheet,
        "value": value,
        "formula": formula,
        "deps": [],
        "ranges": [],
        "external": False,
    }
    if depth >= max_depth:
        node["truncated"] = True
        return node
    refs, ranges, has_external = parse_refs(formula, sheet)
    node["external"] = has_external
    next_path = set(path)
    next_path.add(key)
    seen: set[str] = set()
    for ref_sheet, ref_cell in refs:
        dep_key = f"{ref_sheet}!{ref_cell}"
        if dep_key in seen:
            continue
        seen.add(dep_key)
        node["deps"].append(_trace_node(entry, ref_sheet, ref_cell, next_path, depth + 1, max_depth))
    for range_sheet, range_ref in ranges:
        node["ranges"].append(
            {
                "sheet": range_sheet,
                "range": range_ref,
                "headers": _get_range_headers(entry, range_sheet, range_ref),
            }
        )
        for coord in _expand_range_formula_cells(entry, range_sheet, range_ref):
            dep_key = f"{range_sheet}!{coord}"
            if dep_key in seen:
                continue
            seen.add(dep_key)
            node["deps"].append(_trace_node(entry, range_sheet, coord, next_path, depth + 1, max_depth))
    return node


def _build_ref_index(fid: str, force: bool = False) -> dict[str, list[tuple[str, str]]]:
    if not force and fid in ref_index_cache:
        return ref_index_cache[fid]
    entry = _file_or_404(fid)
    index: dict[str, list[tuple[str, str]]] = {}
    for sheet_name in entry["wb_f"].sheetnames:
        ws = entry["wb_f"][sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                formula = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
                if not formula:
                    continue
                refs, ranges, _ = parse_refs(formula, sheet_name)
                for ref_sheet, ref_cell in refs:
                    index.setdefault(f"{ref_sheet}!{ref_cell}", []).append((sheet_name, cell.coordinate))
                for ref_sheet, range_ref in ranges:
                    index.setdefault(f"RANGE:{ref_sheet}!{range_ref}", []).append((sheet_name, cell.coordinate))
    ref_index_cache[fid] = index
    return index


def _trace_up(entry: dict[str, Any], fid: str, sheet: str, cell: str, path: set[str] | None = None) -> dict[str, Any]:
    path = path or set()
    key = f"{sheet}!{cell}"
    if key in path:
        return {"cell": cell, "sheet": sheet, "value": "[circular]", "formula": None, "deps": [], "ranges": [], "external": False}
    index = _build_ref_index(fid)
    parents = list(index.get(key, []))
    for idx_key, refs in index.items():
        if not idx_key.startswith("RANGE:"):
            continue
        target = idx_key[len("RANGE:"):]
        target_sheet, range_ref = target.split("!", 1)
        if target_sheet == sheet and _cell_in_range(cell, range_ref):
            parents.extend(refs)
    node = _trace_node(entry, sheet, cell, set(), 0, MAX_TRACE_DEPTH)
    next_path = set(path)
    next_path.add(key)
    deduped = list(dict.fromkeys(parents))
    node["deps"] = [_trace_up(entry, fid, parent_sheet, parent_cell, next_path) for parent_sheet, parent_cell in deduped]
    return node


def _sheet_extent(ws) -> tuple[int, int]:
    max_row = min(ws.max_row or 1, 5000)
    max_col = min(ws.max_column or 1, 500)
    if max_row * max_col <= 200_000:
        return max_row, max_col
    for row in range(max_row, max(1, max_row - 5000), -1):
        for col in range(max_col, max(1, max_col - 500), -1):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return row, col
    return max_row, max_col


def _sheet_data(entry: dict[str, Any], fid: str, sheet: str) -> dict[str, Any]:
    cache_key = f"{fid}:{sheet}"
    if cache_key in sheet_cache:
        return sheet_cache[cache_key]
    ws_f = entry["wb_f"][sheet]
    ws_v = entry["wb_v"][sheet]
    max_row, max_col = _sheet_extent(ws_f)
    headers = [_col_to_letter(c) for c in range(1, max_col + 1)]
    rows: list[list[dict[str, Any]]] = []
    for row_num in range(1, max_row + 1):
        row_payload: list[dict[str, Any]] = []
        for col_num in range(1, max_col + 1):
            coord = f"{_col_to_letter(col_num)}{row_num}"
            formula = _cell_formula(ws_f, coord)
            value = _cell_value(ws_v, coord)
            if formula is None and value == "":
                continue
            cell_payload: dict[str, Any] = {"r": coord, "v": value, "f": formula}
            style = _extract_cell_style(ws_f[coord])
            if style:
                cell_payload["s"] = style
            if formula:
                meta = _compute_cell_label(entry, sheet, coord)
                if meta and meta != f"{sheet}!{coord}":
                    cell_payload["m"] = meta
            row_payload.append(cell_payload)
        rows.append(row_payload)
    payload = {"headers": headers, "rows": rows}
    sheet_cache[cache_key] = payload
    return payload


def _detect_tables(entry: dict[str, Any], fid: str, sheet: str) -> list[dict[str, Any]]:
    cache_key = f"{fid}:{sheet}"
    if cache_key in tables_cache:
        return tables_cache[cache_key]
    persisted = UPLOADS_DIR / fid / f"tables_{sheet}.json"
    if persisted.exists():
        try:
            tables_cache[cache_key] = json.loads(persisted.read_text())
            return tables_cache[cache_key]
        except Exception:
            pass
    ws = entry["wb_f"][sheet]
    occupied = {(cell.row, cell.column) for row in ws.iter_rows() for cell in row if cell.value not in (None, "")}
    seen: set[tuple[int, int]] = set()
    tables: list[dict[str, Any]] = []
    for start in list(occupied):
        if start in seen:
            continue
        queue = [start]
        component: list[tuple[int, int]] = []
        seen.add(start)
        while queue:
            row, col = queue.pop()
            component.append((row, col))
            for nxt in ((row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)):
                if nxt in occupied and nxt not in seen:
                    seen.add(nxt)
                    queue.append(nxt)
        rows = [item[0] for item in component]
        cols = [item[1] for item in component]
        min_row, max_row = min(rows), max(rows)
        min_col, max_col = min(cols), max(cols)
        row_count = max_row - min_row + 1
        col_count = max_col - min_col + 1
        if len(component) < 8 or row_count < 3 or col_count < 2:
            continue
        formulas = numbers = texts = 0
        preview: list[list[str]] = []
        headers: list[str] = []
        for col in range(min_col, max_col + 1):
            head = ws.cell(row=min_row, column=col).value
            if isinstance(head, str) and head.strip() and not _is_numericish(head):
                headers.append(head.strip())
        for row in range(min_row, max_row + 1):
            preview_row: list[str] = []
            for col in range(min_col, max_col + 1):
                raw = ws.cell(row=row, column=col).value
                if isinstance(raw, str) and raw.startswith("="):
                    formulas += 1
                elif isinstance(raw, (int, float)):
                    numbers += 1
                elif raw not in (None, ""):
                    texts += 1
                if row <= min_row + 3:
                    preview_row.append("" if raw is None else str(raw))
            if preview_row:
                preview.append(preview_row)
        tables.append(
            {
                "range": f"{_col_to_letter(min_col)}{min_row}:{_col_to_letter(max_col)}{max_row}",
                "top_left": f"{_col_to_letter(min_col)}{min_row}",
                "rows": row_count,
                "cols": col_count,
                "cells": len(component),
                "formulas": formulas,
                "numbers": numbers,
                "texts": texts,
                "headers": headers,
                "has_header": bool(headers),
                "preview": preview[:4],
            }
        )
    tables.sort(key=lambda item: (range_boundaries(item["range"])[1], range_boundaries(item["range"])[0]))
    tables = tables[:50]
    tables_cache[cache_key] = tables
    _atomic_write_json(persisted, tables)
    return tables


def _formula_ref_count(formula: str | None, sheet: str) -> int:
    refs, ranges, _ = parse_refs(formula, sheet)
    return len(refs) + len(ranges)


def _build_range_bounds(fid: str) -> dict[str, list[tuple[int, int, int, int]]]:
    index = _build_ref_index(fid)
    out: dict[str, list[tuple[int, int, int, int]]] = {}
    for key in index:
        if not key.startswith("RANGE:"):
            continue
        target = key[len("RANGE:"):]
        sheet, range_ref = target.split("!", 1)
        try:
            out.setdefault(sheet, []).append(range_boundaries(range_ref))
        except Exception:
            continue
    return out


def _cell_in_any_range(cell: str, bounds: list[tuple[int, int, int, int]]) -> bool:
    parsed = _parse_cell_ref(cell)
    if not parsed:
        return False
    col, row = parsed
    for min_col, min_row, max_col, max_row in bounds:
        if min_col <= col <= max_col and min_row <= row <= max_row:
            return True
    return False


def _get_referenced_set(fid: str) -> set[str]:
    if fid in _referenced_cache:
        return _referenced_cache[fid]
    index = _build_ref_index(fid)
    referenced: set[str] = {key for key in index if not key.startswith("RANGE:")}
    for key in index:
        if not key.startswith("RANGE:"):
            continue
        target = key[len("RANGE:"):]
        sheet_name, range_ref = target.split("!", 1)
        try:
            min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        except Exception:
            continue
        scanned = 0
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                referenced.add(f"{sheet_name}!{_col_to_letter(col)}{row}")
                scanned += 1
                if scanned >= 10_000:
                    break
            if scanned >= 10_000:
                break
    _referenced_cache[fid] = referenced
    return referenced


def _trace_to_text(node: dict[str, Any], depth: int = 0) -> str:
    indent = "    " * depth
    meta = f" | Context: {node['meta']}" if node.get("meta") else ""
    line = f"{indent}- {node['sheet']}!{node['cell']} | Formula: {node.get('formula')} | Value: {node.get('value')}{meta}"
    range_lines = [
        f"{indent}    Range: {rng['sheet']}!{rng['range']}" + (f" | Headers: {', '.join(rng.get('headers', []))}" if rng.get("headers") else "")
        for rng in node.get("ranges", [])
    ]
    child_lines = [_trace_to_text(dep, depth + 1) for dep in node.get("deps", [])]
    return "\n".join([line, *range_lines, *child_lines])


def _collect_sheets(node: dict[str, Any], out: set[str] | None = None) -> list[str]:
    out = out or set()
    out.add(node["sheet"])
    for dep in node.get("deps", []):
        _collect_sheets(dep, out)
    return sorted(out)


def _build_formula_text(node: dict[str, Any], depth: int = 0) -> str:
    indent = "  " * depth
    label = node.get("meta") or f"{node['sheet']}!{node['cell']}"
    line = f"{indent}{label} = {node.get('formula') or node.get('value')}"
    children = [_build_formula_text(dep, depth + 1) for dep in node.get("deps", [])]
    return "\n".join([line, *children])


def _count_nodes(node: dict[str, Any]) -> tuple[int, int, set[str]]:
    total = 1
    formulas = 1 if node.get("formula") else 0
    sheets = {node["sheet"]}
    for dep in node.get("deps", []):
        child_total, child_formulas, child_sheets = _count_nodes(dep)
        total += child_total
        formulas += child_formulas
        sheets.update(child_sheets)
    return total, formulas, sheets


def _read_prompt(name: str) -> str:
    return (PROMPTS_DIR / name).read_text()


def _llm_client() -> LLMClient:
    return LLMClient(
        api_env=get_config_value("EFT_API_ENV") or "test",
        app_name=get_config_value("APP_NAME") or "formulalens",
        aws_region=get_config_value("AWS_REGION") or "us-east-1",
    )


def _stream_llm(prompt_name: str, user_text: str, model: str, temperature: float, max_tokens: int) -> Generator[str, None, dict[str, Any] | None]:
    client = _llm_client()
    final_usage: dict[str, Any] | None = None
    stream = client.stream_openai(
        model=model,
        messages=[{"role": "user", "content": user_text}],
        system_prompt=_read_prompt(prompt_name),
        max_tokens=max_tokens,
        temperature=temperature,
    )
    for chunk in stream:
        if isinstance(chunk, dict):
            final_usage = chunk
        else:
            yield chunk
    return final_usage


def _make_cache_key(sheet: str, cell: str, task_type: str) -> str:
    return f"{sheet}!{cell}:{task_type}"


def _run_llm_task(task: LLMTask, chain_fn, chain_kwargs: dict[str, Any]) -> None:
    try:
        for chunk in chain_fn(**chain_kwargs):
            if task.cancel_flag.is_set():
                task.status = "cancelled"
                return
            task.chunks.append(chunk)
            task.full_text += chunk
        task.status = "done"
        if task.cache_key and task.file_id:
            explanation_cache[f"{task.file_id}:{task.cache_key}"] = task.full_text
            _save_explanations(task.file_id)
    except Exception as exc:
        task.status = "error"
        task.error_msg = str(exc)


def _start_or_reconnect_task(
    *,
    task_type: str,
    cache_key: str,
    file_id: str,
    regenerate: bool,
    chain_fn,
    chain_kwargs: dict[str, Any],
) -> dict[str, Any]:
    if file_id and cache_key and not regenerate:
        cache_hit = explanation_cache.get(f"{file_id}:{cache_key}")
        if cache_hit:
            return {"cached": True, "text": cache_hit}
    with _task_lock:
        if cache_key and not regenerate and cache_key in task_by_cache_key:
            existing_id = task_by_cache_key[cache_key]
            existing = task_store.get(existing_id)
            if existing and existing.status == "running":
                return {"task_id": existing.task_id}
        task_id = str(uuid.uuid4())
        task = LLMTask(task_id=task_id, task_type=task_type, cache_key=cache_key, file_id=file_id)
        task_store[task_id] = task
        if cache_key:
            task_by_cache_key[cache_key] = task_id
    thread = threading.Thread(target=_run_llm_task, args=(task, chain_fn, chain_kwargs), daemon=True)
    thread.start()
    return {"task_id": task_id}


def _cleanup_old_tasks() -> None:
    while True:
        time.sleep(300)
        cutoff = time.time() - 600
        with _task_lock:
            expired = [
                task_id
                for task_id, task in task_store.items()
                if task.status != "running" and task.created_at < cutoff
            ]
            for task_id in expired:
                task = task_store.pop(task_id, None)
                if not task:
                    continue
                if task.cache_key and task_by_cache_key.get(task.cache_key) == task_id:
                    task_by_cache_key.pop(task.cache_key, None)


def _stream_technical_explanation(trace_text: str, model: str, max_tokens: int, label: str) -> Generator[str, None, None]:
    user_text = (
        f"Here is the full dependency tree for the metric '{label}':\n\n"
        f"{trace_text}\n\nPlease explain this formula in plain English."
    )
    yield from _stream_llm("explain_formula.txt", user_text, model, 0.2, max_tokens)


def _stream_business_summary(trace_text: str, model: str, max_tokens: int, label: str) -> Generator[str, None, None]:
    user_text = (
        f"Here is the full dependency tree for the metric '{label}':\n\n"
        f"{trace_text}\n\nExplain this metric from a business perspective."
    )
    yield from _stream_llm("business_summary.txt", user_text, model, 0.2, max_tokens)


def _stream_formula_reconstruction(trace_text: str, model: str, max_tokens: int, label: str) -> Generator[str, None, None]:
    user_text = (
        f"Here is the full dependency tree for the metric '{label}':\n\n"
        f"{trace_text}\n\nShow me how to reconstruct this formula from scratch and how it could be rewritten more cleanly."
    )
    yield from _stream_llm("formula_reconstruction.txt", user_text, model, 0.2, max_tokens)


def _stream_formula_snapshot(trace_text: str, model: str, max_tokens: int, label: str) -> Generator[str, None, None]:
    user_text = (
        f"Here is the full dependency tree for the metric '{label}':\n\n"
        f"{trace_text}\n\nGenerate a concise formula snapshot."
    )
    yield from _stream_llm("formula_snapshot.txt", user_text, model, 0.1, max_tokens)


def _infer_persona(message: str, mode: str) -> str:
    if mode in {"excel", "data", "business"}:
        return mode
    lowered = message.lower()
    if any(hint in lowered for hint in _PERSONA_BUSINESS_HINTS):
        return "business"
    if any(hint in lowered for hint in _PERSONA_DATA_HINTS):
        return "data"
    if any(hint in lowered for hint in _PERSONA_EXCEL_HINTS):
        return "excel"
    return "excel"


def _persona_overlay(persona: str) -> str:
    if persona == "data":
        return "Data analyst lens. Look for trends, distributions, outliers, quantify everything. End with **Insight:**."
    if persona == "business":
        return "Executive QBR format. Lead with headline, then **Drivers** and **Watch-outs/Risks**. Translate formulas to business meaning."
    return "Answer like a senior FP&A modeler. Lead with formula mechanics, cite exact cell refs in backticks, suggest cleaner rewrites."


def _search_cells_by_query(entry: dict[str, Any], query: str) -> list[dict[str, Any]]:
    lowered = query.lower()
    out: list[dict[str, Any]] = []
    for sheet_name in entry["wb_v"].sheetnames:
        ws = entry["wb_v"][sheet_name]
        max_row = min(ws.max_row or 1, 200)
        max_col = min(ws.max_column or 1, 50)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                raw = ws.cell(row=row, column=col).value
                if raw is None:
                    continue
                text = str(raw)
                if lowered in text.lower():
                    out.append({"sheet": sheet_name, "cell": f"{_col_to_letter(col)}{row}", "value": text})
                    if len(out) >= 30:
                        return out
    return out


def _is_blocked(message: str) -> bool:
    lowered = message.lower()
    if any(keyword in lowered for keyword in _INJECTION_KEYWORDS):
        return True
    if any(target in lowered for target in _INJECTION_TARGETS) and any(word in lowered for word in ("reveal", "print", "show", "dump")):
        return True
    if re.search(r"\byou are now\b|\bact as\b|\bpretend to be\b", lowered):
        return True
    return any(pattern.search(message) for pattern in _HARM_PHRASES)


def _stream_chat_response(user_message: str, context: str, history: list[dict], model: str, max_tokens: int, persona: str) -> Generator[str, None, None]:
    system_prompt = _read_prompt("chat.txt") + "\n\n" + _persona_overlay(persona)
    messages: list[dict[str, str]] = []
    for item in history[-10:]:
        role = item.get("role")
        content = str(item.get("content", ""))
        if role in {"user", "assistant"} and content:
            messages.append({"role": role, "content": content})
    messages.append({"role": "user", "content": f"Workbook context:\n{context}\n\nUser request:\n{user_message}"})
    client = _llm_client()
    stream = client.stream_openai(
        model=model,
        messages=messages,
        system_prompt=system_prompt,
        max_tokens=max_tokens,
        temperature=0.3,
    )
    for chunk in stream:
        if not isinstance(chunk, dict):
            yield chunk


def _load_registry() -> None:
    _boot_status.update({"stage": "registry", "detail": "Loading workbook registry", "files_total": 0, "files_loaded": 0, "ready": False})
    if not REGISTRY_PATH.exists():
        _boot_status.update({"stage": "ready", "detail": "No saved workbooks", "ready": True})
        _registry_ready.set()
        return
    try:
        data = json.loads(REGISTRY_PATH.read_text())
    except Exception:
        _boot_status.update({"stage": "ready", "detail": "Registry unreadable, starting fresh", "ready": True})
        _registry_ready.set()
        return
    if not isinstance(data, dict):
        _boot_status.update({"stage": "ready", "detail": "Registry format invalid", "ready": True})
        _registry_ready.set()
        return
    items = [(fid, item) for fid, item in data.items() if FILE_ID_RE.fullmatch(fid)]
    _boot_status["files_total"] = len(items)
    for fid, item in items:
        folder = UPLOADS_DIR / fid
        workbook_path = Path(str(item.get("path", "")))
        if not workbook_path.exists() and folder.exists():
            matches = list(folder.glob("*.xlsx"))
            if matches:
                workbook_path = matches[0]
        if not workbook_path.exists():
            continue
        try:
            wb_f = load_workbook(workbook_path, data_only=False)
            wb_v = load_workbook(workbook_path, data_only=True)
            store[fid] = {
                "path": str(workbook_path),
                "filename": item.get("filename") or workbook_path.name,
                "wb_f": wb_f,
                "wb_v": wb_v,
                "sheets": wb_f.sheetnames,
            }
            _boot_status.update({"stage": "indexing", "detail": f"Loading {workbook_path.name}"})
            _build_ref_index(fid, force=True)
            _load_explanations(fid)
            _boot_status["files_loaded"] += 1
        except Exception:
            continue
    _boot_status.update({"stage": "ready", "detail": "Workbook registry loaded", "ready": True})
    _registry_ready.set()


@app.middleware("http")
async def wait_for_registry(request: Request, call_next):
    path = request.url.path
    skip = path in ("/ping", "/api/upload", "/api/status")
    if not skip and path.startswith("/api") and not _registry_ready.is_set():
        _registry_ready.wait(timeout=180)
        if not _registry_ready.is_set():
            return JSONResponse({"detail": "Server still loading workbook data"}, status_code=503)
    return await call_next(request)


@app.on_event("startup")
def startup() -> None:
    threading.Thread(target=_load_registry, daemon=True).start()
    threading.Thread(target=_cleanup_old_tasks, daemon=True).start()


@app.get("/ping")
def ping() -> dict[str, Any]:
    return {"status": "ok", "ready": _registry_ready.is_set()}


@app.get("/api/status")
def api_status() -> dict[str, Any]:
    return dict(_boot_status)


@app.post("/api/upload")
async def upload(file: UploadFile = File(...)):
    async def event_stream():
        filename = _sanitize_filename(file.filename or "workbook.xlsx")
        if not filename.lower().endswith(".xlsx"):
            yield _sse({"error": "Only .xlsx files are supported"})
            return
        raw = await file.read()
        if len(raw) > MAX_UPLOAD_SIZE:
            yield _sse({"error": "File exceeds 200 MB limit"})
            return
        yield _sse({"progress": "Reading formulas..."})
        fid = uuid.uuid4().hex[:12]
        folder = UPLOADS_DIR / fid
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / filename
        path.write_bytes(raw)
        try:
            wb_f = load_workbook(path, data_only=False)
            yield _sse({"progress": f"Found {len(wb_f.sheetnames)} sheets — reading values..."})
            wb_v = load_workbook(path, data_only=True)
        except Exception as exc:
            shutil.rmtree(folder, ignore_errors=True)
            yield _sse({"error": f"Failed to parse workbook: {exc}"})
            return
        store[fid] = {
            "path": str(path),
            "filename": filename,
            "wb_f": wb_f,
            "wb_v": wb_v,
            "sheets": wb_f.sheetnames,
        }
        yield _sse({"progress": f"Building formula index across {len(wb_f.sheetnames)} sheets...", "indexing": {"current": 0, "total": len(wb_f.sheetnames), "sheet": ""}})

        def background_index():
            _build_ref_index(fid, force=True)

        threading.Thread(target=background_index, daemon=True).start()
        _save_registry()
        yield _sse({"done": True, "file_id": fid, "filename": filename, "sheets": wb_f.sheetnames})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.get("/api/files")
def files() -> list[dict[str, Any]]:
    return [{"file_id": fid, "filename": entry["filename"], "sheets": entry["sheets"]} for fid, entry in store.items()]


@app.get("/api/files/{fid}")
def file_detail(fid: str) -> dict[str, Any]:
    entry = _file_or_404(fid)
    return {"file_id": fid, "filename": entry["filename"], "sheets": entry["sheets"]}


@app.delete("/api/files/{fid}")
def delete_file(fid: str) -> dict[str, bool]:
    store.pop(fid, None)
    ref_index_cache.pop(fid, None)
    _referenced_cache.pop(fid, None)
    for key in [key for key in list(sheet_cache) if key.startswith(f"{fid}:")]:
        sheet_cache.pop(key, None)
    for key in [key for key in list(tables_cache) if key.startswith(f"{fid}:")]:
        tables_cache.pop(key, None)
    for key in [key for key in list(explanation_cache) if key.startswith(f"{fid}:")]:
        explanation_cache.pop(key, None)
    shutil.rmtree(UPLOADS_DIR / fid, ignore_errors=True)
    _save_registry()
    return {"ok": True}


@app.get("/api/sheet/{fid}/{sheet:path}")
def get_sheet(fid: str, sheet: str) -> dict[str, Any]:
    entry = _file_or_404(fid)
    sheet_name, _, _ = _sheet_or_404(entry, sheet)
    return _sheet_data(entry, fid, sheet_name)


@app.get("/api/sheet-stream/{fid}/{sheet:path}")
def get_sheet_stream(fid: str, sheet: str):
    entry = _file_or_404(fid)
    sheet_name, ws_f, _ = _sheet_or_404(entry, sheet)

    def event_stream():
        max_row, max_col = _sheet_extent(ws_f)
        yield _sse({"progress": f"Reading {max_row} rows x {max_col} columns"})
        step = max(1, max_row // 5)
        for row in range(step, max_row + 1, step):
            pct = min(100, int((row / max_row) * 100))
            yield _sse({"progress": f"Processing row {row} of {max_row} ({pct}%)"})
        data = _sheet_data(entry, fid, sheet_name)
        yield _sse({"done": True, "data": data})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/reload/{fid}")
def reload_workbook(fid: str, sheet: str | None = None) -> dict[str, Any]:
    lock = _get_file_lock(fid)
    with lock:
        _reload_file_from_disk(fid)
        cleared = _clear_file_caches(fid, sheet)
    return {"ok": True, "cleared": cleared}


@app.post("/api/trace")
def trace(req: TraceReq) -> dict[str, Any]:
    entry = _file_or_404(req.file_id)
    trace_node = _trace_node(entry, req.sheet, req.cell.upper(), set(), 0, req.max_depth)
    return {"trace_tree": _add_meta_to_trace(entry, trace_node)}


@app.post("/api/trace-up")
def trace_up(req: TraceReq) -> dict[str, Any]:
    entry = _file_or_404(req.file_id)
    trace_node = _trace_up(entry, req.file_id, req.sheet, req.cell.upper(), set())
    return {"trace_tree": _add_meta_to_trace(entry, trace_node)}


@app.post("/api/table-trace")
def table_trace(req: TableTraceReq) -> dict[str, Any]:
    entry = _file_or_404(req.file_id)
    if req.sheet not in entry["wb_f"].sheetnames:
        raise HTTPException(status_code=404, detail="sheet not found")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(req.range)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid range format") from exc
    ws = entry["wb_f"][req.sheet]
    metrics: list[dict[str, Any]] = []
    cache: dict[str, dict[str, Any]] = {}
    total_formulas = 0
    dep_keys: set[str] = set()
    for row in range(min_row, max_row + 1):
        label = None
        cells: list[dict[str, Any]] = []
        for col in range(min_col, max_col + 1):
            coord = f"{_col_to_letter(col)}{row}"
            raw = ws[coord].value
            if isinstance(raw, str) and raw.startswith("="):
                total_formulas += 1
                if coord not in cache:
                    cache[coord] = _add_meta_to_trace(entry, _trace_node(entry, req.sheet, coord, set(), 0, req.max_depth))
                cells.append(cache[coord])
            elif label is None and isinstance(raw, str) and raw.strip() and not _is_numericish(raw):
                label = raw.strip()
        if cells:
            metrics.append({"label": label or f"Metric {row}", "cells": cells})
            for cell in cells:
                for dep in cell.get("deps", []):
                    dep_keys.add(f"{dep['sheet']}!{dep['cell']}")
    return {"metrics": metrics, "total_formulas": total_formulas, "total_deps": len(dep_keys)}


@app.get("/api/tables/{fid}/{sheet:path}")
def tables(fid: str, sheet: str) -> list[dict[str, Any]]:
    entry = _file_or_404(fid)
    sheet_name, _, _ = _sheet_or_404(entry, sheet)
    return _detect_tables(entry, fid, sheet_name)


@app.put("/api/tables/{fid}/{sheet:path}")
def save_tables(fid: str, sheet: str, payload: SaveTablesReq) -> dict[str, bool]:
    sheet_name = unquote(sheet)
    tables_cache[f"{fid}:{sheet_name}"] = payload.tables
    _atomic_write_json(UPLOADS_DIR / fid / f"tables_{sheet_name}.json", payload.tables)
    return {"ok": True}


@app.get("/api/task/{task_id}")
def task_status(task_id: str) -> dict[str, Any]:
    task = task_store.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="task not found")
    return {"task_id": task.task_id, "status": task.status, "chunk_count": len(task.chunks), "full_text": task.full_text}


@app.get("/api/task/{task_id}/stream")
def task_stream(task_id: str, offset: int = 0):
    task = task_store.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="task not found")

    def event_stream():
        cursor = max(0, offset)
        while True:
            current = task_store.get(task_id)
            if not current:
                yield _sse({"error": "Task expired"})
                return
            while cursor < len(current.chunks):
                yield _sse({"text": current.chunks[cursor], "offset": cursor})
                cursor += 1
            if current.status == "done":
                yield _sse({"done": True, "status": "done", "full_text": current.full_text})
                return
            if current.status == "error":
                yield _sse({"error": current.error_msg or "Task failed", "status": "error"})
                return
            if current.status == "cancelled":
                yield _sse({"done": True, "status": "cancelled", "full_text": current.full_text})
                return
            time.sleep(0.15)

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/task/{task_id}/cancel")
def cancel_task(task_id: str) -> dict[str, str]:
    task = task_store.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="task not found")
    task.cancel_flag.set()
    return {"status": "cancelling"}


def _start_explain_task(req: ExplainReq, task_type: str, max_tokens: int, chain_fn):
    trace = dict(req.trace)
    label = trace.get("meta") or f"{trace.get('sheet', req.sheet)}!{trace.get('cell', req.cell)}"
    trace_text = _trace_to_text(trace)
    cache_key = _make_cache_key(req.sheet, req.cell, task_type) if req.file_id and req.sheet and req.cell else ""
    return _start_or_reconnect_task(
        task_type=task_type,
        cache_key=cache_key,
        file_id=req.file_id,
        regenerate=req.regenerate,
        chain_fn=chain_fn,
        chain_kwargs={"trace_text": trace_text, "model": req.model or DEFAULT_MODEL, "max_tokens": max_tokens, "label": label},
    )


@app.post("/api/explain")
def explain(req: ExplainReq) -> dict[str, Any]:
    return _start_explain_task(req, "analyst", 4096, _stream_technical_explanation)


@app.post("/api/business-summary")
def business_summary(req: ExplainReq) -> dict[str, Any]:
    return _start_explain_task(req, "business", 2000, _stream_business_summary)


@app.post("/api/reconstruct")
def reconstruct(req: ExplainReq) -> dict[str, Any]:
    return _start_explain_task(req, "blueprint", 4096, _stream_formula_reconstruction)


@app.post("/api/snapshot")
def snapshot(req: ExplainReq) -> dict[str, Any]:
    return _start_explain_task(req, "snapshot", 800, _stream_formula_snapshot)


@app.get("/api/explanations/{fid}/{sheet:path}/{cell}")
def get_cached_explanations(fid: str, sheet: str, cell: str) -> dict[str, str]:
    sheet_name = unquote(sheet)
    return {
        "analyst": explanation_cache.get(f"{fid}:{_make_cache_key(sheet_name, cell, 'analyst')}", ""),
        "business": explanation_cache.get(f"{fid}:{_make_cache_key(sheet_name, cell, 'business')}", ""),
    }


@app.post("/api/table-explain-batch")
def table_explain_batch(req: BatchExplainReq):
    def event_stream():
        for idx, metric in enumerate(req.metrics):
            trace = metric.get("trace", {})
            label = metric.get("label") or trace.get("meta") or "Metric"
            trace_text = _trace_to_text(trace)
            for kind, fn, max_tokens in (
                ("analyst", _stream_technical_explanation, 4096),
                ("business", _stream_business_summary, 2000),
            ):
                yield _sse({"metric_index": idx, "type": kind, "status": "start"})
                full_text = ""
                for chunk in fn(trace_text=trace_text, model=req.model, max_tokens=max_tokens, label=label):
                    full_text += chunk
                    yield _sse({"metric_index": idx, "type": kind, "text": chunk})
                yield _sse({"metric_index": idx, "type": kind, "status": "done", "full_text": full_text})
        yield _sse({"all_done": True})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/top-metrics/explain-all")
def top_metrics_explain_all(req: SummaryExplainReq):
    def event_stream():
        for idx, metric in enumerate(req.metrics):
            trace = metric.get("trace", {})
            label = trace.get("meta") or f"{trace.get('sheet', '')}!{trace.get('cell', '')}"
            trace_text = _trace_to_text(trace)
            for kind, fn, max_tokens in (
                ("analyst", _stream_technical_explanation, 4096),
                ("business", _stream_business_summary, 2000),
                ("blueprint", _stream_formula_reconstruction, 4096),
            ):
                yield _sse({"metric_index": idx, "type": kind, "status": "start"})
                full_text = ""
                for chunk in fn(trace_text=trace_text, model=req.model, max_tokens=max_tokens, label=label):
                    full_text += chunk
                    yield _sse({"metric_index": idx, "type": kind, "text": chunk})
                yield _sse({"metric_index": idx, "type": kind, "status": "done", "full_text": full_text})
        yield _sse({"all_done": True})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/optimize")
def optimize(req: OptimizeReq):
    trace = dict(req.trace)
    total_nodes, formula_count, sheets = _count_nodes(trace)
    trace_text = _trace_to_text(trace)
    user_text = (
        f"Metric: {req.label or trace.get('meta') or 'Metric'}\n"
        f"Stats: {total_nodes} nodes, {formula_count} formulas, {len(sheets)} sheets ({', '.join(sorted(sheets))})\n\n"
        f"Full dependency tree:\n\n{trace_text}\n\nAnalyze this formula tree and determine if it can be optimized."
    )

    def event_stream():
        buffer = ""
        try:
            for chunk in _stream_llm("optimize_formula.txt", user_text, req.model, 0.3, 4000):
                buffer += chunk
                yield _sse({"text": chunk})
        except Exception as exc:
            yield _sse({"error": str(exc)})
            return
        match = re.search(r"```json\s*(\{.*\})\s*```", buffer, re.S)
        if not match:
            match = re.search(r"(\{.*\})\s*$", buffer, re.S)
        result = {"verdict": "keep", "reason": "Could not parse optimization result."}
        if match:
            try:
                result = json.loads(match.group(1))
            except Exception:
                pass
        if result.get("verdict") not in {"keep", "optimize"}:
            result = {"verdict": "keep", "reason": "Optimization verdict was invalid."}
        if result.get("verdict") == "optimize":
            tree = result.setdefault("optimized_tree", trace)
            tree.setdefault("ranges", [])
            tree.setdefault("deps", [])
        yield _sse({"result": result, "done": True})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.post("/api/chat")
def chat(req: ChatReq):
    if len(req.message) > 50_000:
        raise HTTPException(status_code=400, detail="Message exceeds 50,000 characters")
    if _is_blocked(req.message):
        def blocked_stream():
            yield _sse({"text": "I can help analyze workbook data, formulas, charts, and cell edits, but I can’t follow that request."})
            yield _sse({"done": True})
        return StreamingResponse(blocked_stream(), media_type="text/event-stream")
    entry = _file_or_404(req.file_id)
    persona = _infer_persona(req.message, req.mode)
    context_parts = []
    if req.sheet and req.sheet in entry["wb_f"].sheetnames:
        context_parts.append(f"Active sheet: {req.sheet}")
    if req.focus_cells:
        context_parts.append(f"Focus cells: {', '.join(req.focus_cells)}")
    if req.selected_tables:
        context_parts.append(f"Selected tables: {', '.join(req.selected_tables)}")
    matches = _search_cells_by_query(entry, req.message)
    if matches:
        context_parts.append("Relevant cells:")
        for match in matches[:10]:
            context_parts.append(f"- {match['sheet']}!{match['cell']} = {match['value']}")
    context = "\n".join(context_parts) or "Workbook loaded. Use the workbook context only."

    def event_stream():
        try:
            yield _sse({"status": "Gathering workbook context"})
            yield _sse({"status": f"Using {persona} lens"})
            for chunk in _stream_chat_response(req.message, context, req.history, req.model, 4096, persona):
                yield _sse({"text": chunk})
            yield _sse({"done": True})
        except Exception as exc:
            yield _sse({"error": str(exc)})

    return StreamingResponse(event_stream(), media_type="text/event-stream")


def _expand_cells(cells: list[str]) -> list[str]:
    expanded: list[str] = []
    for item in cells:
        if ":" not in item:
            expanded.append(item.upper())
            continue
        try:
            min_col, min_row, max_col, max_row = range_boundaries(item.upper())
        except Exception:
            continue
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                expanded.append(f"{_col_to_letter(col)}{row}")
    return list(dict.fromkeys(expanded))


@app.post("/api/edit-cells")
def edit_cells(req: EditCellsReq) -> dict[str, Any]:
    if len(req.edits) > 200:
        raise HTTPException(status_code=400, detail="Maximum 200 edits per request")
    entry = _file_or_404(req.file_id)
    if req.sheet not in entry["wb_f"].sheetnames:
        raise HTTPException(status_code=404, detail="sheet not found")
    lock = _get_file_lock(req.file_id)
    results = []
    with lock:
        ws_f = entry["wb_f"][req.sheet]
        ws_v = entry["wb_v"][req.sheet]
        for edit in req.edits:
            coord = edit.cell.upper()
            if edit.formula is not None and edit.value is not None:
                raise HTTPException(status_code=400, detail=f"{coord} cannot include both value and formula")
            if edit.formula is not None:
                formula = edit.formula if edit.formula.startswith("=") else f"={edit.formula}"
                ws_f[coord] = formula
                ws_v[coord] = None
                results.append({"cell": coord, "status": "formula"})
            else:
                ws_f[coord] = edit.value
                ws_v[coord] = edit.value
                results.append({"cell": coord, "status": "value"})
        entry["wb_f"].save(entry["path"])
        _reload_file_from_disk(req.file_id)
        _clear_file_caches(req.file_id, req.sheet)
    return {"ok": True, "results": results}


@app.post("/api/format-cells")
def format_cells(req: FormatCellsReq) -> dict[str, Any]:
    entry = _file_or_404(req.file_id)
    if req.sheet not in entry["wb_f"].sheetnames:
        raise HTTPException(status_code=404, detail="sheet not found")
    expanded = _expand_cells(req.cells)
    lock = _get_file_lock(req.file_id)
    with lock:
        for wb_key in ("wb_f", "wb_v"):
            ws = entry[wb_key][req.sheet]
            for coord in expanded:
                cell = ws[coord]
                if req.format.fill is not None:
                    if req.format.fill:
                        cell.fill = PatternFill(fill_type="solid", fgColor=req.format.fill.replace("#", "").upper())
                    else:
                        cell.fill = PatternFill(fill_type=None)
                font_kwargs = {
                    "name": cell.font.name,
                    "size": cell.font.sz,
                    "bold": req.format.bold if req.format.bold is not None else cell.font.bold,
                    "italic": req.format.italic if req.format.italic is not None else cell.font.italic,
                }
                if req.format.font_color is not None and req.format.font_color:
                    font_kwargs["color"] = req.format.font_color.replace("#", "").upper()
                elif req.format.font_color == "":
                    font_kwargs["color"] = None
                else:
                    font_kwargs["color"] = getattr(cell.font.color, "rgb", None)
                cell.font = Font(**font_kwargs)
                if req.format.number_format is not None:
                    cell.number_format = "General" if req.format.number_format == "" else req.format.number_format
        entry["wb_f"].save(entry["path"])
        _reload_file_from_disk(req.file_id)
        _clear_file_caches(req.file_id, req.sheet)
    return {"ok": True, "cells_updated": len(expanded)}


def _is_band_empty(ws, row0: int, col0: int, height: int, width: int) -> bool:
    for row in range(row0, row0 + height):
        for col in range(col0, col0 + width):
            if ws.cell(row=row, column=col).value not in (None, ""):
                return False
    return True


def _find_chart_placement(ws, band_h: int, band_w: int, near_range: str | None) -> tuple[int, int]:
    if near_range:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(near_range)
            candidates = [
                (min_row, max_col + 2),
                (max_row + 2, min_col),
            ]
            candidates.extend((min_row, max_col + offset) for offset in range(4, 22, 2))
            candidates.extend((max_row + offset, min_col) for offset in range(4, 32, 2))
            candidates.append((max_row + 2, min_col))
            for row0, col0 in candidates:
                if _is_band_empty(ws, row0, col0, band_h, band_w):
                    return row0, col0
        except Exception:
            pass
    return (ws.max_row or 1) + 3, 1


@app.post("/api/insert-chart")
def insert_chart(req: InsertChartReq) -> dict[str, Any]:
    entry = _file_or_404(req.file_id)
    if req.sheet not in entry["wb_f"].sheetnames:
        raise HTTPException(status_code=404, detail="sheet not found")
    series = req.chart_spec.get("series", [])
    total_points = sum(len(item.get("data", [])) for item in series)
    if total_points > 5000:
        raise HTTPException(status_code=400, detail="Chart exceeds 5,000 data points")
    if len(series) > 50:
        raise HTTPException(status_code=400, detail="Chart exceeds 50 series")
    lock = _get_file_lock(req.file_id)
    with lock:
        ws_f = entry["wb_f"][req.sheet]
        ws_v = entry["wb_v"][req.sheet]
        labels = [str(item.get("label", "")) for item in (series[0].get("data", []) if series else [])]
        band_h = max(3, len(labels) + 1)
        band_w = max(2, len(series) + 1)
        row0, col0 = _find_chart_placement(ws_v, band_h, band_w, req.near_range)
        ws_f.cell(row=row0, column=col0, value="Category")
        ws_v.cell(row=row0, column=col0, value="Category")
        for idx, item in enumerate(series, start=1):
            ws_f.cell(row=row0, column=col0 + idx, value=item.get("name", f"Series {idx}"))
            ws_v.cell(row=row0, column=col0 + idx, value=item.get("name", f"Series {idx}"))
        for ridx, label in enumerate(labels, start=1):
            ws_f.cell(row=row0 + ridx, column=col0, value=label)
            ws_v.cell(row=row0 + ridx, column=col0, value=label)
            for sidx, item in enumerate(series, start=1):
                point = item.get("data", [])[ridx - 1]
                value = point.get("value", 0)
                ws_f.cell(row=row0 + ridx, column=col0 + sidx, value=value)
                ws_v.cell(row=row0 + ridx, column=col0 + sidx, value=value)
        chart_type = (req.chart_spec.get("type") or "bar").lower()
        chart = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "area": AreaChart,
            "scatter": ScatterChart,
        }.get(chart_type, BarChart)()
        chart.title = req.chart_spec.get("title") or "Chart"
        chart.width = 18
        chart.height = 12
        data_ref = Reference(ws_f, min_col=col0 + 1, min_row=row0, max_col=col0 + len(series), max_row=row0 + len(labels))
        cats_ref = Reference(ws_f, min_col=col0, min_row=row0 + 1, max_row=row0 + len(labels))
        if isinstance(chart, PieChart):
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
        elif isinstance(chart, ScatterChart):
            for sidx in range(len(series)):
                xvalues = Reference(ws_f, min_col=col0, min_row=row0 + 1, max_row=row0 + len(labels))
                yvalues = Reference(ws_f, min_col=col0 + 1 + sidx, min_row=row0, max_row=row0 + len(labels))
                chart.series.append(Series(yvalues, xvalues, title_from_data=True))
        else:
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
        anchor_col = col0 + band_w + 2
        anchor = f"{_col_to_letter(anchor_col)}{row0}"
        ws_f.add_chart(chart, anchor)
        entry["wb_f"].save(entry["path"])
        _reload_file_from_disk(req.file_id)
        _clear_file_caches(req.file_id, req.sheet)
    data_range = f"{_col_to_letter(col0)}{row0}:{_col_to_letter(col0 + band_w - 1)}{row0 + band_h - 1}"
    return {"ok": True, "data_range": data_range, "chart_anchor": anchor}


@app.get("/api/download/{fid}")
def download(fid: str):
    entry = _file_or_404(fid)
    src = Path(entry["path"])
    tmp = Path(tempfile.gettempdir()) / f"{fid}-{src.name}"
    shutil.copy2(src, tmp)
    return FileResponse(tmp, filename=entry["filename"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/top-metrics/{fid}")
def top_metrics(fid: str, sheets: str | None = None, min_refs: int = 2) -> dict[str, Any]:
    entry = _file_or_404(fid)
    selected_sheets = [sheet for sheet in (sheets.split(",") if sheets else entry["sheets"]) if sheet in entry["wb_f"].sheetnames]
    index = _build_ref_index(fid)
    direct_refs = {key for key in index if not key.startswith("RANGE:")}
    range_bounds = _build_range_bounds(fid)
    metrics: list[dict[str, Any]] = []
    seen_labels: set[str] = set()
    for sheet_name in selected_sheets:
        ws_f = entry["wb_f"][sheet_name]
        ws_v = entry["wb_v"][sheet_name]
        max_row, max_col = _sheet_extent(ws_f)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                coord = f"{_col_to_letter(col)}{row}"
                formula = _cell_formula(ws_f, coord)
                if not formula:
                    continue
                if f"{sheet_name}!{coord}" in direct_refs:
                    continue
                if _cell_in_any_range(coord, range_bounds.get(sheet_name, [])):
                    continue
                if _formula_ref_count(formula, sheet_name) < max(1, min_refs):
                    continue
                label = _get_cell_label(entry, sheet_name, coord)
                if not label or label == f"{sheet_name}!{coord}" or label.upper() in {"N/M", "NM", "N/A"} or label in seen_labels:
                    continue
                seen_labels.add(label)
                metrics.append(
                    {
                        "sheet": sheet_name,
                        "cell": coord,
                        "label": label,
                        "formula": formula,
                        "value": _cell_value(ws_v, coord),
                    }
                )
                if len(metrics) >= 200:
                    return {"metrics": metrics, "total": len(metrics)}
    return {"metrics": metrics, "total": len(metrics)}


@app.post("/api/top-metrics/{fid}/trace/{sheet:path}/{cell}")
def top_metric_trace(fid: str, sheet: str, cell: str, max_depth: int = 5) -> dict[str, Any]:
    entry = _file_or_404(fid)
    sheet_name = unquote(sheet)
    trace = _add_meta_to_trace(entry, _trace_node(entry, sheet_name, cell.upper(), set(), 0, max_depth))
    return {"trace": trace, "sheets_involved": _collect_sheets(trace), "formula_text": _build_formula_text(trace)}
