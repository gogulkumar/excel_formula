#!/usr/bin/env python3
"""
Comprehensive LLM connection test script for CalcSense.

Tests all LLM integration points:
  1. LLMClient initialization & configuration
  2. Mock mode: call_openai, stream_openai, call_claude, stream_claude
  3. Live mode (OpenAI direct via OPENAI_API_KEY): real API call if key is set
  4. All chain functions: explain, business, reconstruct, snapshot, optimize, chat
  5. Missing prompt file detection
  6. End-to-end via backend API (explain, business-summary, chat, optimize)

Usage:
  .venv/bin/python scripts/test_llm_connections.py
  .venv/bin/python scripts/test_llm_connections.py --live    # also tests real OpenAI API
  .venv/bin/python scripts/test_llm_connections.py --backend http://127.0.0.1:8010  # also tests via API
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
import time
import traceback
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
APP_ROOT = REPO_ROOT / "app"
sys.path.insert(0, str(APP_ROOT))

# ─── Results tracking ────────────────────────────────────────────────────────

_results: list[dict[str, Any]] = []


def _record(test_name: str, passed: bool, detail: str = "") -> None:
    status = "PASS" if passed else "FAIL"
    icon = "✅" if passed else "❌"
    _results.append({"test": test_name, "passed": passed, "detail": detail})
    print(f"  {icon}  {status}  {test_name}" + (f" — {detail}" if detail else ""))


# ─── Test 1: Config Loader ───────────────────────────────────────────────────

def test_config_loader() -> None:
    print("\n━━━ 1. Config Loader ━━━")
    try:
        from config_loader import get_config_value, get_config_bool
        mode = get_config_value("EFT_LLM_MODE")
        _record("config_loader import", True, f"EFT_LLM_MODE={mode}")
    except Exception as exc:
        _record("config_loader import", False, str(exc))


# ─── Test 2: LLMClient Initialization ────────────────────────────────────────

def test_llm_client_init() -> None:
    print("\n━━━ 2. LLMClient Initialization ━━━")
    try:
        from llm_client import LLMClient
        client = LLMClient()
        _record("LLMClient instantiation", True)
        _record("is_mock_mode", True, f"is_mock_mode={client.is_mock_mode}")
        _record("openai_url", True, f"openai_url='{client.openai_url}' (empty is OK in mock)")
        _record("bedrock_url", True, f"bedrock_url='{client.bedrock_url}' (empty is OK in mock)")
        _record("runtime", True, f"runtime='{client.runtime}'")
        _record("model_mode", True, f"model_mode='{client.model_mode}'")
    except Exception as exc:
        _record("LLMClient instantiation", False, str(exc))


# ─── Test 3: Mock Mode ─ call_openai ─────────────────────────────────────────

def test_mock_call_openai() -> None:
    print("\n━━━ 3. call_openai ━━━")
    try:
        from llm_client import LLMClient
        client = LLMClient()
        mode_label = "mock" if client.is_mock_mode else "live"
        messages = [{"role": "user", "content": "Explain =SUM(A1:A5)"}]
        result = client.call_openai(model="gpt-4", messages=messages)
        assert isinstance(result, dict), "Expected dict result"
        assert "response_text" in result, "Missing response_text"
        assert len(result["response_text"]) > 0, "Empty response_text"
        _record(f"call_openai ({mode_label})", True, f"response length={len(result['response_text'])}")
    except Exception as exc:
        _record("call_openai", False, str(exc))


# ─── Test 4: Mock Mode ─ stream_openai ───────────────────────────────────────

def test_mock_stream_openai() -> None:
    print("\n━━━ 4. Mock Mode — stream_openai ━━━")
    try:
        from llm_client import LLMClient
        client = LLMClient()
        messages = [{"role": "user", "content": "Explain =SUM(A1:A5)"}]
        chunks = list(client.stream_openai(model="gpt-4", messages=messages))
        string_chunks = [c for c in chunks if isinstance(c, str)]
        dict_chunks = [c for c in chunks if isinstance(c, dict)]
        assert len(string_chunks) >= 1, "Expected at least 1 text chunk"
        assert len(dict_chunks) >= 1, "Expected final usage dict"
        _record("stream_openai (mock)", True, f"text_chunks={len(string_chunks)}, usage_dicts={len(dict_chunks)}")
    except Exception as exc:
        _record("stream_openai (mock)", False, str(exc))


# ─── Test 5: Mock Mode ─ call_claude ─────────────────────────────────────────

def test_mock_call_claude() -> None:
    print("\n━━━ 5. Mock Mode — call_claude ━━━")
    try:
        from llm_client import LLMClient
        client = LLMClient()
        messages = [{"role": "user", "content": "Summarize this metric"}]
        result = client.call_claude(model="claude-3-sonnet", messages=messages)
        assert isinstance(result, dict), "Expected dict result"
        assert "response_text" in result, "Missing response_text"
        assert len(result["response_text"]) > 0, "Empty response_text"
        _record("call_claude (mock)", True, f"response length={len(result['response_text'])}")
    except Exception as exc:
        _record("call_claude (mock)", False, str(exc))


# ─── Test 6: Mock Mode ─ stream_claude ───────────────────────────────────────

def test_mock_stream_claude() -> None:
    print("\n━━━ 6. Mock Mode — stream_claude ━━━")
    try:
        from llm_client import LLMClient
        client = LLMClient()
        messages = [{"role": "user", "content": "Summarize this metric"}]
        chunks = list(client.stream_claude(model="claude-3-sonnet", messages=messages))
        string_chunks = [c for c in chunks if isinstance(c, str)]
        dict_chunks = [c for c in chunks if isinstance(c, dict)]
        assert len(string_chunks) >= 1, "Expected at least 1 text chunk"
        assert len(dict_chunks) >= 1, "Expected final usage dict"
        _record("stream_claude (mock)", True, f"text_chunks={len(string_chunks)}, usage_dicts={len(dict_chunks)}")
    except Exception as exc:
        _record("stream_claude (mock)", False, str(exc))


# ─── Test 7: Mock Prompt Kinds ───────────────────────────────────────────────

def test_mock_prompt_kinds() -> None:
    print("\n━━━ 7. Mock Prompt Kinds (explain, business, optimize) ━━━")
    from llm_client import LLMClient
    client = LLMClient()
    messages = [{"role": "user", "content": "Test prompt"}]

    for kind in ("explain", "business", "optimize"):
        try:
            chunks = list(client.stream_openai(model="gpt-4", messages=messages, prompt_kind=kind))
            text_parts = [c for c in chunks if isinstance(c, str)]
            full = "".join(text_parts)
            assert len(full) > 10, f"Mock for '{kind}' too short"
            _record(f"mock prompt_kind='{kind}'", True, f"len={len(full)}")
        except Exception as exc:
            _record(f"mock prompt_kind='{kind}'", False, str(exc))


# ─── Test 8: Prompt Files ────────────────────────────────────────────────────

def test_prompt_files() -> None:
    print("\n━━━ 8. Prompt File Availability ━━━")
    prompts_dir = APP_ROOT / "prompts"

    # Files referenced in the codebase
    required_prompts = {
        "explain_formula.txt": "explain_chain.py",
        "business_summary.txt": "business_chain.py",
        "optimize_formula.txt": "optimize_chain.py",
        "formula_reconstruction.txt": "reconstruction_chain.py",
        "formula_snapshot.txt": "snapshot_chain.py",
        "chat.txt": "chat_chain.py",
    }

    for filename, chain in required_prompts.items():
        path = prompts_dir / filename
        exists = path.exists()
        if exists:
            content = path.read_text().strip()
            _record(f"prompt '{filename}'", bool(content), f"referenced by {chain}, {len(content)} chars")
        else:
            _record(f"prompt '{filename}'", False, f"MISSING — referenced by {chain}")


# ─── Test 9: Chain Functions (Mock Mode) ─────────────────────────────────────

def test_chains_mock() -> None:
    print("\n━━━ 9. Chain Functions (Mock Mode) ━━━")
    # Need to be in the backend context for chains to import
    sys.path.insert(0, str(APP_ROOT / "backend"))

    trace_text = (
        "- Summary!B4 | Formula: =B2-B3 | Value: 160 | Context: Profit\n"
        "    - Summary!B2 | Formula: =Revenue!B5+Revenue!C5 | Value: 860 | Context: Revenue\n"
        "    - Summary!B3 | Formula: =Revenue!D5 | Value: 265 | Context: Expenses"
    )

    chains_to_test = [
        ("stream_technical_explanation", "explain_chain", {"trace_text": trace_text, "model": "gpt-4", "max_tokens": 2000, "label": "Profit"}),
        ("stream_business_summary", "business_chain", {"trace_text": trace_text, "model": "gpt-4", "max_tokens": 2000, "label": "Profit"}),
        ("stream_formula_reconstruction", "reconstruction_chain", {"trace_text": trace_text, "model": "gpt-4", "max_tokens": 2000, "label": "Profit"}),
        ("stream_formula_snapshot", "snapshot_chain", {"trace_text": trace_text, "model": "gpt-4", "max_tokens": 800, "label": "Profit"}),
        (
            "stream_optimization",
            "optimize_chain",
            {"trace_text": trace_text, "label": "Profit", "total_nodes": 3, "formula_count": 3, "sheets_involved": ["Summary", "Revenue"], "model": "gpt-4", "max_tokens": 2000},
        ),
    ]

    for fn_name, module_name, kwargs in chains_to_test:
        try:
            module = __import__(f"chains.{module_name}", fromlist=[fn_name])
            fn = getattr(module, fn_name)
            chunks = list(fn(**kwargs))
            full = "".join(chunks)
            assert len(full) > 0, "Empty output from chain"
            _record(f"chain '{fn_name}'", True, f"output_len={len(full)}")
        except FileNotFoundError as exc:
            _record(f"chain '{fn_name}'", False, f"Missing prompt file: {exc}")
        except Exception as exc:
            _record(f"chain '{fn_name}'", False, str(exc))

    # Chat chain
    try:
        from chains.chat_chain import stream_chat_response, infer_persona
        persona = infer_persona("What is the formula in B4?", "auto")
        _record("infer_persona", True, f"persona='{persona}'")

        chunks = list(stream_chat_response(
            user_message="What is the formula in B4?",
            context="Active sheet: Summary",
            history=[],
            model="gpt-4",
            max_tokens=2000,
            persona=persona,
        ))
        full = "".join(chunks)
        assert len(full) > 0, "Empty chat output"
        _record("chain 'stream_chat_response'", True, f"output_len={len(full)}")
    except FileNotFoundError as exc:
        _record("chain 'stream_chat_response'", False, f"Missing prompt file: {exc}")
    except Exception as exc:
        _record("chain 'stream_chat_response'", False, str(exc))


# ─── Test 10: Live OpenAI API (optional) ─────────────────────────────────────

def test_live_openai() -> None:
    print("\n━━━ 10. Live OpenAI API ━━━")
    api_key = os.environ.get("OPENAI_API_KEY") or ""
    if not api_key:
        # Try from .env
        env_path = APP_ROOT / ".env"
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                if line.startswith("OPENAI_API_KEY="):
                    api_key = line.split("=", 1)[1].strip().strip('"')
                    break

    if not api_key or api_key.startswith("your-"):
        _record("live OpenAI API", False, "No valid OPENAI_API_KEY found — skipping")
        return

    try:
        import httpx
        response = httpx.post(
            "https://api.openai.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": "gpt-4o-mini",
                "messages": [{"role": "user", "content": "Say 'CalcSense LLM test OK' in exactly those words."}],
                "max_tokens": 30,
                "temperature": 0.0,
            },
            timeout=30.0,
        )
        if response.status_code == 200:
            data = response.json()
            text = data["choices"][0]["message"]["content"]
            _record("live OpenAI API call", True, f"response='{text[:80]}'")
        elif response.status_code == 401:
            _record("live OpenAI API call", False, "401 Unauthorized — API key is invalid or expired")
        elif response.status_code == 429:
            _record("live OpenAI API call", False, "429 Rate Limited — key is valid but quota exceeded")
        else:
            _record("live OpenAI API call", False, f"HTTP {response.status_code}: {response.text[:200]}")
    except Exception as exc:
        _record("live OpenAI API call", False, str(exc))


# ─── Test 11: Backend API LLM Endpoints (optional) ───────────────────────────

def test_backend_llm_endpoints(base_url: str) -> None:
    print(f"\n━━━ 11. Backend API LLM Endpoints ({base_url}) ━━━")
    import httpx
    from openpyxl import Workbook

    # Check if the backend is reachable
    try:
        ping = httpx.get(f"{base_url}/ping", timeout=5.0)
        if ping.status_code != 200:
            _record("backend reachable", False, f"HTTP {ping.status_code}")
            return
        _record("backend reachable", True)
    except Exception as exc:
        _record("backend reachable", False, f"Cannot reach backend: {exc}")
        return

    # Upload a test workbook
    with tempfile.TemporaryDirectory() as tmpdir:
        wb_path = Path(tmpdir) / "llm_test.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"], ws["B1"] = "Metric", "Value"
        ws["A2"], ws["B2"] = "Revenue", 1000
        ws["A3"], ws["B3"] = "Costs", 400
        ws["A4"], ws["B4"] = "Profit", "=B2-B3"
        wb.save(wb_path)

        with wb_path.open("rb") as fh:
            resp = httpx.post(
                f"{base_url}/api/upload",
                files={"file": ("llm_test.xlsx", fh, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30.0,
            )
        events = []
        for line in resp.text.splitlines():
            if line.startswith("data: "):
                events.append(json.loads(line[6:]))
        done_event = next((e for e in events if e.get("done")), None)
        if not done_event:
            _record("test workbook upload", False, f"Upload failed: {events}")
            return
        file_id = str(done_event["file_id"])
        _record("test workbook upload", True, f"file_id={file_id}")

    # Trace B4
    trace_resp = httpx.post(
        f"{base_url}/api/trace",
        json={"file_id": file_id, "sheet": "Summary", "cell": "B4"},
        timeout=10.0,
    ).json()
    trace = trace_resp["trace_tree"]
    _record("trace B4", True, f"formula={trace.get('formula')}")

    # Test /api/explain
    try:
        explain_resp = httpx.post(
            f"{base_url}/api/explain",
            json={"trace": trace, "file_id": file_id, "sheet": "Summary", "cell": "B4", "regenerate": True},
            timeout=30.0,
        ).json()
        if explain_resp.get("task_id"):
            task_id = explain_resp["task_id"]
            text = _wait_task(base_url, task_id)
            _record("API /api/explain", bool(text), f"len={len(text)}")
        elif explain_resp.get("cached"):
            _record("API /api/explain", True, "cached hit")
        else:
            _record("API /api/explain", False, str(explain_resp))
    except Exception as exc:
        _record("API /api/explain", False, str(exc))

    # Test /api/business-summary
    try:
        biz_resp = httpx.post(
            f"{base_url}/api/business-summary",
            json={"trace": trace, "file_id": file_id, "sheet": "Summary", "cell": "B4", "regenerate": True},
            timeout=30.0,
        ).json()
        if biz_resp.get("task_id"):
            text = _wait_task(base_url, biz_resp["task_id"])
            _record("API /api/business-summary", bool(text), f"len={len(text)}")
        elif biz_resp.get("cached"):
            _record("API /api/business-summary", True, "cached hit")
        else:
            _record("API /api/business-summary", False, str(biz_resp))
    except Exception as exc:
        _record("API /api/business-summary", False, str(exc))

    # Test /api/chat
    try:
        chat_resp = httpx.post(
            f"{base_url}/api/chat",
            json={"file_id": file_id, "message": "What does B4 represent?", "sheet": "Summary", "history": []},
            timeout=30.0,
        )
        chat_events = []
        for line in chat_resp.text.splitlines():
            if line.startswith("data: "):
                chat_events.append(json.loads(line[6:]))
        text_parts = [e.get("text", "") for e in chat_events if "text" in e]
        full_text = "".join(text_parts)
        done = any(e.get("done") for e in chat_events)
        error = next((e.get("error") for e in chat_events if "error" in e), None)
        if error:
            _record("API /api/chat", False, f"error={error}")
        elif done and full_text:
            _record("API /api/chat", True, f"len={len(full_text)}")
        else:
            _record("API /api/chat", False, f"done={done}, text_len={len(full_text)}")
    except Exception as exc:
        _record("API /api/chat", False, str(exc))

    # Test /api/optimize
    try:
        opt_resp = httpx.post(
            f"{base_url}/api/optimize",
            json={"trace": trace, "label": "Profit"},
            timeout=30.0,
        )
        opt_events = []
        for line in opt_resp.text.splitlines():
            if line.startswith("data: "):
                opt_events.append(json.loads(line[6:]))
        done_event = next((e for e in opt_events if e.get("done")), None)
        error_event = next((e for e in opt_events if "error" in e), None)
        if error_event:
            _record("API /api/optimize", False, f"error={error_event['error']}")
        elif done_event and done_event.get("result"):
            _record("API /api/optimize", True, f"verdict={done_event['result'].get('verdict')}")
        else:
            _record("API /api/optimize", False, f"events={len(opt_events)}")
    except Exception as exc:
        _record("API /api/optimize", False, str(exc))

    # Test /api/explanations cache
    try:
        cache_resp = httpx.get(f"{base_url}/api/explanations/{file_id}/Summary/B4", timeout=10.0).json()
        has_analyst = bool(cache_resp.get("analyst"))
        has_business = bool(cache_resp.get("business"))
        _record("API /api/explanations cache", has_analyst or has_business, f"analyst={has_analyst}, business={has_business}")
    except Exception as exc:
        _record("API /api/explanations cache", False, str(exc))

    # Cleanup
    try:
        httpx.delete(f"{base_url}/api/files/{file_id}", timeout=10.0)
    except Exception:
        pass


def _wait_task(base_url: str, task_id: str, timeout_s: float = 30.0) -> str:
    import httpx
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        payload = httpx.get(f"{base_url}/api/task/{task_id}", timeout=5.0).json()
        if payload["status"] == "done":
            return payload.get("full_text", "")
        if payload["status"] == "error":
            raise RuntimeError(f"task error: {payload}")
        time.sleep(0.3)
    raise RuntimeError("task timed out")


# ─── Summary ─────────────────────────────────────────────────────────────────

def print_summary() -> None:
    print("\n" + "═" * 70)
    print("  SUMMARY")
    print("═" * 70)
    passed = sum(1 for r in _results if r["passed"])
    failed = sum(1 for r in _results if not r["passed"])
    total = len(_results)
    print(f"  Total: {total}   Passed: {passed}   Failed: {failed}")

    if failed:
        print(f"\n  ❌ FAILURES:")
        for r in _results:
            if not r["passed"]:
                print(f"     • {r['test']}: {r['detail']}")
    print("═" * 70)


# ─── Main ────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="CalcSense LLM connection test suite")
    parser.add_argument("--live", action="store_true", help="Also test live OpenAI API calls")
    parser.add_argument("--backend", type=str, default="", help="Backend URL for API endpoint tests (e.g. http://127.0.0.1:8010)")
    args = parser.parse_args()

    print("╔══════════════════════════════════════════════════════════════════╗")
    print("║          CalcSense LLM Connection Test Suite                    ║")
    print("╚══════════════════════════════════════════════════════════════════╝")

    test_config_loader()
    test_llm_client_init()
    test_mock_call_openai()
    test_mock_stream_openai()
    test_mock_call_claude()
    test_mock_stream_claude()
    test_mock_prompt_kinds()
    test_prompt_files()
    test_chains_mock()

    if args.live:
        test_live_openai()

    if args.backend:
        test_backend_llm_endpoints(args.backend)

    print_summary()
    failed = sum(1 for r in _results if not r["passed"])
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
